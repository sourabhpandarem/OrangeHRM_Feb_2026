
import openpyxl

class XLUtils_class:
    @staticmethod
    def get_row_count(file,sheet):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheet]
        return sheet.max_row

    @staticmethod
    def get_column_count(file,sheet):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheet]
        return sheet.max_column

    @staticmethod
    def read_data(file_path,sheet_name,row_num,col_num):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        return sheet.cell(row=row_num,column=col_num).value

    @staticmethod
    def write_data(file_path,sheet_name,row_num,col_num,data):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        sheet.cell(row=row_num,column=col_num).value = data
        workbook.save(file_path)
        workbook.close()

